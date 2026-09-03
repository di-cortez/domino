#!/usr/bin/env python3
"""Generate count heatmaps and an Excel workbook from reward lookup manifests.

Place this file in:
    analysis/reward_lookup_table/

Expected inputs:
    derived/double-three_reward_lookup_manifest.json
    derived/double-four_reward_lookup_manifest.json
    derived/double-five_reward_lookup_manifest.json
    derived/double-six_reward_lookup_manifest.json

Run:
    python3 generate_reward_lookup_count_visuals.py

Dependencies:
    numpy, matplotlib, openpyxl
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

try:
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.colors import LogNorm, to_hex
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Dependência ausente. Em um ambiente virtual, execute:\n"
        "  python3 -m pip install numpy matplotlib openpyxl\n"
        f"Erro original: {exc}"
    ) from exc


RULESET_ORDER = ["double-three", "double-four", "double-five", "double-six"]
DISPLAY_NAMES = {
    "double-three": "Double-three",
    "double-four": "Double-four",
    "double-five": "Double-five",
    "double-six": "Double-six",
}

# Limiar que será útil na etapa seguinte, ao escolher células confiáveis.
# Ele não elimina células dos gráficos/tabelas de counts: apenas é informado
# no terminal ao final da execução.
SIGNIFICANT_FRACTION = 0.005

COLORS = {
    "navy": "0F172A",
    "teal": "0F766E",
    "teal_light": "CCFBF1",
    "slate": "475569",
    "line": "CBD5E1",
    "total": "FEF3C7",
    "white": "FFFFFF",
    "zero": "F1F5F9",
}


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Gera quatro heatmaps de counts e uma planilha com quatro abas."
    )
    parser.add_argument(
        "--derived-dir",
        type=Path,
        default=script_dir / "derived",
        help="Diretório que contém os manifests (default: ./derived).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "count_visualizations",
        help="Diretório de saída (default: ./count_visualizations).",
    )
    return parser.parse_args()


def pt_int(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def load_manifests(derived_dir: Path) -> dict[str, dict[str, Any]]:
    manifests: dict[str, dict[str, Any]] = {}
    for ruleset in RULESET_ORDER:
        path = derived_dir / f"{ruleset}_reward_lookup_manifest.json"
        if not path.is_file():
            raise FileNotFoundError(f"Manifest não encontrado: {path}")
        data = json.loads(path.read_text(encoding="utf-8"))
        if data.get("ruleset_name") != ruleset:
            raise ValueError(
                f"ruleset_name inesperado em {path}: {data.get('ruleset_name')!r}"
            )
        data["_source_path"] = str(path)
        manifests[ruleset] = data
    return manifests


def make_matrix(manifest: dict[str, Any]) -> np.ndarray:
    counts = manifest.get("cell_sample_counts")
    if not isinstance(counts, dict) or not counts:
        raise ValueError(
            f"cell_sample_counts ausente ou vazio em {manifest['_source_path']}"
        )

    parsed: dict[tuple[int, int], int] = {}
    for key, value in counts.items():
        try:
            first, second = map(int, key.split(","))
            parsed[first, second] = int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Chave/count inválido: {key!r}: {value!r}") from exc

    max_first = max(first for first, _ in parsed)
    max_second = max(second for _, second in parsed)
    matrix = np.zeros((max_first, max_second), dtype=np.int64)
    for (first, second), value in parsed.items():
        if first < 1 or second < 1 or value < 0:
            raise ValueError(f"Índice/count inválido: {(first, second)} -> {value}")
        matrix[first - 1, second - 1] = value

    expected = int(manifest["summary"]["decisions"])
    actual = int(matrix.sum())
    if actual != expected:
        raise ValueError(
            f"Soma divergente em {manifest['ruleset_name']}: "
            f"matriz={actual}, summary.decisions={expected}"
        )
    return matrix


def draw_heatmap(
    ruleset: str,
    manifest: dict[str, Any],
    matrix: np.ndarray,
    global_max: int,
    png_path: Path,
    pdf: PdfPages,
) -> None:
    rows, columns = matrix.shape
    figure_width = max(8.2, 0.82 * columns + 3.2)
    figure_height = max(6.2, 0.67 * rows + 3.0)
    figure, axis = plt.subplots(
        figsize=(figure_width, figure_height), constrained_layout=True
    )

    masked = np.ma.masked_where(matrix == 0, matrix)
    color_map = plt.get_cmap("YlOrRd").copy()
    color_map.set_bad("#F1F5F9")
    normalization = LogNorm(vmin=1, vmax=global_max)
    image = axis.imshow(
        masked,
        cmap=color_map,
        norm=normalization,
        aspect="equal",
    )

    axis.set_xticks(np.arange(columns), labels=np.arange(1, columns + 1))
    axis.set_yticks(np.arange(rows), labels=np.arange(1, rows + 1))
    axis.set_xlabel("Peças na mão do oponente (2º índice)", fontsize=11, labelpad=10)
    axis.set_ylabel(
        "Peças na mão do agente neural (1º índice)", fontsize=11, labelpad=10
    )
    axis.tick_params(top=True, bottom=False, labeltop=True, labelbottom=False)

    axis.set_xticks(np.arange(-0.5, columns, 1), minor=True)
    axis.set_yticks(np.arange(-0.5, rows, 1), minor=True)
    axis.grid(which="minor", color="white", linewidth=1.5)
    axis.tick_params(which="minor", bottom=False, left=False)
    for spine in axis.spines.values():
        spine.set_visible(False)

    for row in range(rows):
        for column in range(columns):
            count = int(matrix[row, column])
            if count == 0:
                label = "0"
                text_color = "#94A3B8"
            else:
                label = pt_int(count)
                text_color = "white" if normalization(count) > 0.69 else "#111827"
            axis.text(
                column,
                row,
                label,
                ha="center",
                va="center",
                fontsize=7.7 if max(rows, columns) >= 10 else 9.0,
                color=text_color,
                fontweight="semibold" if count >= 10_000 else "normal",
            )

    summary = manifest["summary"]
    title = f"{DISPLAY_NAMES[ruleset]} - número de amostras por tamanho das mãos"
    subtitle = (
        f"{pt_int(int(summary['games']))} jogos · "
        f"{pt_int(int(summary['decisions']))} decisões · "
        f"{int(summary['cells'])} células observadas"
    )
    axis.set_title(f"{title}\n{subtitle}", fontsize=14, fontweight="bold", pad=18)

    color_bar = figure.colorbar(image, ax=axis, fraction=0.045, pad=0.035)
    color_bar.set_label("Count (escala logarítmica)", rotation=90, labelpad=12)
    ticks = [1, 10, 100, 1_000, 10_000]
    if global_max > 10_000:
        ticks.append(global_max)
    ticks = sorted(set(tick for tick in ticks if tick <= global_max))
    color_bar.set_ticks(ticks)
    color_bar.set_ticklabels([pt_int(int(tick)) for tick in ticks])

    figure.savefig(png_path, dpi=240, bbox_inches="tight", facecolor="white")
    pdf.savefig(figure, bbox_inches="tight", facecolor="white")
    plt.close(figure)


def apply_fill(cell, rgb: str) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=rgb.replace("#", "").upper())


def build_workbook(
    manifests: dict[str, dict[str, Any]],
    matrices: dict[str, np.ndarray],
    png_paths: dict[str, Path],
    global_max: int,
    workbook_path: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    thin = Side(style="thin", color=COLORS["line"])
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    color_map = plt.get_cmap("YlOrRd")
    normalization = LogNorm(vmin=1, vmax=global_max)

    for ruleset in RULESET_ORDER:
        manifest = manifests[ruleset]
        matrix = matrices[ruleset]
        summary = manifest["summary"]
        rows, columns = matrix.shape
        total_column = columns + 2
        total_row = 6 + rows
        title_end_column = max(total_column, 9)

        sheet = workbook.create_sheet(DISPLAY_NAMES[ruleset])
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "B6"

        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=title_end_column
        )
        title_cell = sheet.cell(1, 1)
        title_cell.value = f"{DISPLAY_NAMES[ruleset]} - counts por tamanho das mãos"
        apply_fill(title_cell, COLORS["navy"])
        title_cell.font = Font(color=COLORS["white"], bold=True, size=16)
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=title_end_column
        )
        summary_cell = sheet.cell(2, 1)
        summary_cell.value = (
            f"{pt_int(int(summary['games']))} jogos | "
            f"{pt_int(int(summary['decisions']))} decisões | "
            f"{int(summary['cells'])} células observadas | "
            f"{pt_int(int(summary['neural_wins']))} vitórias / "
            f"{pt_int(int(summary['neural_losses']))} derrotas"
        )
        apply_fill(summary_cell, "E2E8F0")
        summary_cell.font = Font(color=COLORS["slate"], size=10)

        sheet.merge_cells(
            start_row=3, start_column=1, end_row=3, end_column=title_end_column
        )
        note_cell = sheet.cell(3, 1)
        note_cell.value = (
            "Chave i,j: linha i = peças do agente neural; "
            "coluna j = peças do oponente. Zeros indicam ausência de amostras."
        )
        note_cell.font = Font(color=COLORS["slate"], italic=True, size=9)

        header_values = [
            "Agente \\ Oponente",
            *range(1, columns + 1),
            "Total da linha",
        ]
        for column, value in enumerate(header_values, start=1):
            cell = sheet.cell(5, column, value)
            apply_fill(cell, COLORS["teal"])
            cell.font = Font(color=COLORS["white"], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = grid_border
        sheet.row_dimensions[5].height = 34

        for row_index in range(rows):
            excel_row = 6 + row_index
            label_cell = sheet.cell(excel_row, 1, row_index + 1)
            apply_fill(label_cell, COLORS["teal_light"])
            label_cell.font = Font(color=COLORS["navy"], bold=True)
            label_cell.alignment = Alignment(horizontal="center")
            label_cell.border = grid_border

            for column_index in range(columns):
                excel_column = 2 + column_index
                count = int(matrix[row_index, column_index])
                cell = sheet.cell(excel_row, excel_column, count)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
                cell.border = grid_border
                if count == 0:
                    apply_fill(cell, COLORS["zero"])
                    cell.font = Font(color="94A3B8")
                else:
                    rgb = to_hex(color_map(normalization(count)), keep_alpha=False)
                    apply_fill(cell, rgb)
                    cell.font = Font(
                        color=COLORS["white"]
                        if normalization(count) > 0.69
                        else COLORS["navy"],
                        bold=count >= 10_000,
                    )

            row_total_cell = sheet.cell(excel_row, total_column, int(matrix[row_index].sum()))
            apply_fill(row_total_cell, COLORS["total"])
            row_total_cell.font = Font(color=COLORS["navy"], bold=True)
            row_total_cell.number_format = "#,##0"
            row_total_cell.alignment = Alignment(horizontal="right")
            row_total_cell.border = grid_border

        total_label = sheet.cell(total_row, 1, "Total da coluna")
        apply_fill(total_label, COLORS["total"])
        total_label.font = Font(color=COLORS["navy"], bold=True)
        total_label.alignment = Alignment(horizontal="right")
        total_label.border = grid_border

        column_totals = matrix.sum(axis=0)
        for column_index, total in enumerate(column_totals, start=2):
            cell = sheet.cell(total_row, column_index, int(total))
            apply_fill(cell, COLORS["total"])
            cell.font = Font(color=COLORS["navy"], bold=True)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            cell.border = grid_border

        grand_total = sheet.cell(total_row, total_column, int(matrix.sum()))
        apply_fill(grand_total, COLORS["total"])
        grand_total.font = Font(color=COLORS["navy"], bold=True)
        grand_total.number_format = "#,##0"
        grand_total.alignment = Alignment(horizontal="right")
        grand_total.border = grid_border

        sheet.column_dimensions["A"].width = 24
        for column in range(2, columns + 2):
            sheet.column_dimensions[get_column_letter(column)].width = 11
        sheet.column_dimensions[get_column_letter(total_column)].width = 17
        sheet.auto_filter.ref = (
            f"A5:{get_column_letter(total_column)}{total_row - 1}"
        )

        image = ExcelImage(str(png_paths[ruleset]))
        image.width = max(720, columns * 72)
        image.height = max(540, rows * 58)
        image.anchor = f"{get_column_letter(total_column + 3)}5"
        sheet.add_image(image)

        threshold = math.ceil(SIGNIFICANT_FRACTION * int(summary["decisions"]))
        sheet.cell(total_row + 2, 1, "Limiar de 0,5%")
        sheet.cell(total_row + 2, 2, threshold)
        sheet.cell(total_row + 2, 1).font = Font(color=COLORS["slate"], italic=True)
        sheet.cell(total_row + 2, 2).font = Font(color=COLORS["slate"], italic=True)
        sheet.cell(total_row + 2, 2).number_format = "#,##0"

        # Ao imprimir/exportar a planilha, cada tabela ocupa uma página em
        # paisagem. O heatmap embutido permanece disponível à direita na aba,
        # mas o PDF dedicado já é a versão apropriada para impressão.
        sheet.print_area = (
            f"A1:{get_column_letter(total_column)}{total_row + 2}"
        )
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.35
        sheet.page_margins.bottom = 0.35

    workbook.save(workbook_path)


def main() -> None:
    args = parse_args()
    derived_dir = args.derived_dir.resolve()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    manifests = load_manifests(derived_dir)
    matrices = {
        ruleset: make_matrix(manifests[ruleset]) for ruleset in RULESET_ORDER
    }
    global_max = max(int(matrix.max()) for matrix in matrices.values())

    png_paths = {
        ruleset: output_dir / f"{ruleset}_counts_heatmap.png"
        for ruleset in RULESET_ORDER
    }
    pdf_path = output_dir / "heatmaps_counts_double_three_a_six.pdf"
    with PdfPages(pdf_path) as pdf:
        metadata = pdf.infodict()
        metadata["Title"] = "Counts por tamanho das mãos - double-three a double-six"
        metadata["Subject"] = "Reward lookup table sample counts"
        for ruleset in RULESET_ORDER:
            draw_heatmap(
                ruleset,
                manifests[ruleset],
                matrices[ruleset],
                global_max,
                png_paths[ruleset],
                pdf,
            )

    workbook_path = output_dir / "domino_counts_heatmaps_e_tabelas.xlsx"
    build_workbook(
        manifests,
        matrices,
        png_paths,
        global_max,
        workbook_path,
    )

    print("Arquivos gerados:")
    print(f"  {pdf_path}")
    print(f"  {workbook_path}")
    for ruleset in RULESET_ORDER:
        print(f"  {png_paths[ruleset]}")

    print("\nVerificação e limiar de 0,5% por ruleset:")
    for ruleset in RULESET_ORDER:
        decisions = int(manifests[ruleset]["summary"]["decisions"])
        threshold = math.ceil(SIGNIFICANT_FRACTION * decisions)
        print(
            f"  {ruleset}: soma={int(matrices[ruleset].sum())}; "
            f"summary.decisions={decisions}; limiar={threshold}"
        )


if __name__ == "__main__":
    main()
