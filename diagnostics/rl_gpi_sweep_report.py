"""Build validated reports for the RL games-per-iteration sweep.

The report is intentionally reconstructed from the experiment directory.  It
does not depend on the in-memory sweep driver, which makes ``--report-only``
and resumed experiments produce the same CSV, JSON, workbook, and plots as an
uninterrupted run.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
from pathlib import Path
import statistics
import tempfile
from typing import Any, Iterable


SCHEMA_VERSION = "2.0"
REPORT_CSV_FILES = {
    "runs": "gpi_sweep_runs.csv",
    "aggregates": "gpi_sweep_aggregate.csv",
    "rankings": "gpi_sweep_ranking.csv",
    "autotune": "gpi_sweep_autotune.csv",
    "pairwise_deltas": "gpi_sweep_pairwise_deltas.csv",
}
AUTOTUNE_FIELDS = (
    "run_key", "critic", "seed", "games_per_iteration", "attempt",
    "workers", "duration_s", "games_per_second", "marginal_gain",
    "passed", "failure_reason", "planned_games", "reused_games",
    "discarded_games",
)
RANKING_FIELDS = (
    "critic", "games_per_iteration", "quality_rank", "throughput_rank",
    "pareto_efficient", "one_se_eligible", "recommended",
    "evidence_status", "best_quality_gpi", "best_efficiency_gpi",
    "one_se_threshold", "score_rate_recommended_gpi",
    "score_rate_changes_recommendation", "rule",
)
PAIRWISE_DELTA_FIELDS = (
    "critic", "games_per_iteration",
    "heuristic_win_rate_delta_vs_gpi40",
    "heuristic_score_rate_delta_vs_gpi40",
    "heuristic_win_rate_delta_vs_best", "best_quality_gpi",
)


def file_sha256(path: Path | str) -> str:
    """Return the SHA-256 digest of *path* without loading it all at once."""
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _load_json(path: Path, *, required: bool = False) -> dict[str, Any] | None:
    try:
        with open(path, "r", encoding="utf-8") as stream:
            value = json.load(stream)
    except FileNotFoundError:
        if required:
            raise
        return None
    if not isinstance(value, dict):
        raise ValueError(f"Expected a JSON object in {path}.")
    return value


def _json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if hasattr(value, "item"):
        return value.item()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def _atomic_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.tmp-", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as stream:
            stream.write(text)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_json(path: Path, value: Any) -> None:
    _atomic_text(
        path,
        json.dumps(value, indent=2, ensure_ascii=False, default=_json_default) + "\n",
    )


def _fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for field in row:
            if field not in seen:
                seen.add(field)
                fields.append(field)
    return fields


def _atomic_csv(path: Path, rows: list[dict[str, Any]], fields: Iterable[str] | None = None) -> None:
    fields = list(fields or _fieldnames(rows) or ("status",))
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.tmp-", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.get(key) for key in fields})
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _load_metrics(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    try:
        with open(path, "r", encoding="utf-8") as stream:
            for line_number, line in enumerate(stream, 1):
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"Non-object metric at {path}:{line_number}.")
                rows.append(value)
    except FileNotFoundError:
        pass
    return rows


def _mean(values: Iterable[Any]) -> float | None:
    numbers = [float(value) for value in values if value is not None]
    return statistics.fmean(numbers) if numbers else None


def _std(values: Iterable[Any]) -> float | None:
    numbers = [float(value) for value in values if value is not None]
    if not numbers:
        return None
    return statistics.stdev(numbers) if len(numbers) > 1 else 0.0


def _median(values: Iterable[Any]) -> float | None:
    numbers = [float(value) for value in values if value is not None]
    return statistics.median(numbers) if numbers else None


def _percentile(values: Iterable[Any], percentile: float) -> float | None:
    numbers = sorted(float(value) for value in values if value is not None)
    if not numbers:
        return None
    if len(numbers) == 1:
        return numbers[0]
    position = (len(numbers) - 1) * percentile
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return numbers[lower]
    fraction = position - lower
    return numbers[lower] * (1.0 - fraction) + numbers[upper] * fraction


def _duration_text(seconds: Any) -> str | None:
    if seconds is None:
        return None
    seconds = float(seconds)
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours >= 1:
        return f"{int(hours)}h {int(minutes):02d}m {seconds:05.2f}s"
    if minutes >= 1:
        return f"{int(minutes)}m {seconds:05.2f}s"
    return f"{seconds:.3f}s"


def _diagnostic_columns(prefix: str, summary: dict[str, Any] | None) -> dict[str, Any]:
    if not summary:
        return {}
    counts = summary.get("counts", {})
    rates = summary.get("rates", {})
    parallel = summary.get("parallel", {})
    choices = summary.get("choice_opportunities", {})
    games = int(summary.get("game_count", 0))
    wins = int(counts.get("win", 0))
    draws = int(counts.get("draw", 0))
    duration = summary.get("duration_s")
    by_position = summary.get("by_position", {})
    ci = summary.get("win_ci95", (None, None))
    return {
        f"{prefix}_games": games,
        f"{prefix}_duration_s": duration,
        f"{prefix}_games_per_second": games / duration if duration else None,
        f"{prefix}_workers_requested": parallel.get("requested_workers"),
        f"{prefix}_workers_initial": parallel.get("initial_workers"),
        f"{prefix}_workers_final": parallel.get("final_workers"),
        f"{prefix}_fallback_count": parallel.get("fallback_count"),
        f"{prefix}_safety_capped": parallel.get("safety_capped"),
        f"{prefix}_wins": wins,
        f"{prefix}_draws": draws,
        f"{prefix}_losses": int(counts.get("loss", 0)),
        f"{prefix}_win_rate": rates.get("win"),
        f"{prefix}_draw_rate": rates.get("draw"),
        f"{prefix}_loss_rate": rates.get("loss"),
        f"{prefix}_score_rate": (wins + 0.5 * draws) / games if games else None,
        f"{prefix}_win_ci95_low": ci[0] if len(ci) > 0 else None,
        f"{prefix}_win_ci95_high": ci[1] if len(ci) > 1 else None,
        f"{prefix}_win_rate_player0": by_position.get("0", {}).get("win_rate"),
        f"{prefix}_win_rate_player1": by_position.get("1", {}).get("win_rate"),
        f"{prefix}_mean_turns": summary.get("mean_turns"),
        f"{prefix}_std_turns": summary.get("std_turns"),
        f"{prefix}_mean_agent_remaining_pips": summary.get("mean_agent_remaining_pips"),
        f"{prefix}_mean_opponent_remaining_pips": summary.get("mean_opponent_remaining_pips"),
        f"{prefix}_real_decision_turns": choices.get("real_decision_turns"),
        f"{prefix}_real_decision_rate": choices.get("real_decision_rate"),
        f"{prefix}_forced_tile_turns": choices.get("forced_tile_turns"),
        f"{prefix}_forced_draws": choices.get("forced_draws"),
        f"{prefix}_forced_passes": choices.get("forced_passes"),
        f"{prefix}_choice_histogram": json.dumps(
            choices.get("choice_histogram", {}), sort_keys=True
        ),
    }


def _validate_diagnostic(
    run_dir: Path,
    opponent: str,
    expected_games: int,
    model_sha256: str | None,
    warnings: list[str],
) -> dict[str, Any] | None:
    directory = run_dir / "diagnostics" / f"vs_{opponent}"
    summary = _load_json(directory / "summary.json")
    config = _load_json(directory / "diagnostic_config.json")
    games_path = directory / "games.csv"
    if summary is None:
        return None
    if int(summary.get("game_count", -1)) != int(expected_games):
        warnings.append(
            f"{run_dir.name} vs {opponent}: expected {expected_games} games, "
            f"found {summary.get('game_count')}."
        )
        return None
    try:
        with open(games_path, "r", encoding="utf-8", newline="") as stream:
            row_count = sum(1 for _ in csv.reader(stream)) - 1
    except OSError:
        return None
    if row_count != expected_games:
        warnings.append(
            f"{run_dir.name} vs {opponent}: games.csv has {row_count}/{expected_games} rows."
        )
        return None
    if config is not None and model_sha256 is not None:
        if config.get("model_sha256") != model_sha256:
            warnings.append(f"{run_dir.name} vs {opponent}: diagnostic model hash mismatch.")
            return None
    return summary


def build_run_rows(experiment_dir: Path | str) -> tuple[list[dict[str, Any]], list[str]]:
    """Read and validate one row for every planned sweep point."""
    experiment_dir = Path(experiment_dir)
    manifest = _load_json(experiment_dir / "experiment_manifest.json", required=True)
    plan = manifest.get("plan")
    if not isinstance(plan, list):
        raise ValueError("experiment_manifest.json does not contain a plan list.")
    experiment_config = manifest.get("configuration", {})
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    sl_hashes: dict[str, set[str]] = {}

    for point in sorted(plan, key=lambda item: int(item["execution_order"])):
        run_dir = experiment_dir / "runs" / point["run_key"]
        config = _load_json(run_dir / "run_config.json") or {}
        summary = _load_json(run_dir / "training_summary.json") or {}
        final_model = _load_json(run_dir / "final_model.json") or {}
        metrics = _load_metrics(run_dir / "training_metrics.jsonl")
        status = point.get("status", "planned")
        model_path_value = final_model.get("model_path")
        model_path = Path(model_path_value) if model_path_value else None
        model_hash = final_model.get("model_sha256")
        if model_path is not None and model_path.exists() and model_hash:
            if file_sha256(model_path) != model_hash:
                warnings.append(f"{point['run_key']}: final model hash mismatch.")
                status = "failed"
        sl_hash = config.get("sl_sha256") or manifest.get("sl_checkpoint", {}).get("sha256")
        sl_hashes.setdefault(str(point["critic"]), set()).add(str(sl_hash))

        total_games = int(point["total_training_games"])
        training_wall = summary.get("training_wall_s")
        total_samples = sum(int(item.get("decision_sample_count", 0)) for item in metrics)
        rollout_wall = sum(float(item.get("rollout_duration_s", 0.0)) for item in metrics)
        update_wall = sum(float(item.get("update_duration_s", 0.0)) for item in metrics)
        grad_norms = [item.get("grad_norm") for item in metrics]
        entropies = [item.get("entropy") for item in metrics if item.get("entropy") is not None]
        reward_means = [item.get("reward_mean") for item in metrics]
        reward_stds = [item.get("reward_std") for item in metrics]
        value_losses = [item.get("value_loss") for item in metrics if item.get("value_loss") is not None]
        clipped_count = sum(bool(item.get("grad_clipped")) for item in metrics)
        ppo_rows = [item for item in metrics if item.get("epochs_completed") is not None]
        completed_iterations = len({int(item["iteration"]) for item in metrics if "iteration" in item})
        parallel = summary.get("parallel", {})
        autotune = summary.get("autotune", {})

        heuristic = _validate_diagnostic(
            run_dir, "heuristic", int(config.get("diagnostic_games", experiment_config.get("diagnostic_games", 0))), model_hash, warnings
        )
        random_summary = _validate_diagnostic(
            run_dir, "random", int(config.get("diagnostic_games", experiment_config.get("diagnostic_games", 0))), model_hash, warnings
        )
        if status == "complete" and (heuristic is None or random_summary is None):
            warnings.append(f"{point['run_key']}: marked complete with an invalid diagnostic.")
            status = "incomplete"

        row: dict[str, Any] = {
            "experiment_id": manifest.get("experiment_id"),
            "run_key": point["run_key"],
            "status": status,
            "execution_order": point["execution_order"],
            "critic": point["critic"],
            "seed": point["seed"],
            "games_per_iteration": point["games_per_iteration"],
            "iterations": point["iterations"],
            "completed_metric_iterations": completed_iterations,
            "total_training_games": total_games,
            "learning_rate": config.get("learning_rate", experiment_config.get("learning_rate")),
            "gamma": config.get("gamma", experiment_config.get("gamma")),
            "value_coef": config.get("value_coef") if point["critic"] == "on" else None,
            "entropy_coef": config.get("entropy_coef", experiment_config.get("entropy_coef")),
            "reward_schema": config.get("reward_schema", experiment_config.get("reward_schema")),
            "normalize_advantages": config.get("normalize_advantages", experiment_config.get("normalize_advantages")),
            "clip_grad_norm": config.get("clip_grad_norm", experiment_config.get("clip_grad_norm")),
            "training_opponent": config.get("training_opponent", experiment_config.get("training_opponent")),
            "rl_training_algorithm": summary.get(
                "rl_training_algorithm",
                config.get("rl_training_algorithm"),
            ),
            "pool_refresh_games": config.get(
                "pool_refresh_games",
                experiment_config.get("pool_refresh_games"),
            ),
            "legacy_pool_interval_iterations": config.get(
                "pool_interval",
                experiment_config.get("pool_interval"),
            ),
            "max_pool_size": config.get("max_pool_size", experiment_config.get("max_pool_size")),
            "checkpoint_count": config.get("checkpoint_count", experiment_config.get("checkpoint_count")),
            "sl_checkpoint": config.get("sl_weights_resolved"),
            "sl_sha256": sl_hash,
            "model_path": str(model_path) if model_path else None,
            "model_sha256": model_hash,
            "git_commit": config.get("git_commit", manifest.get("environment", {}).get("git_commit")),
            "device_requested": config.get("device", experiment_config.get("device")),
            "device_selected": summary.get("device"),
            "training_wall_s": training_wall,
            "training_wall_text": _duration_text(training_wall),
            "training_internal_s": summary.get("duration_s"),
            "rollout_wall_s": rollout_wall if metrics else summary.get("total_rollout_duration_s"),
            "update_wall_s": update_wall if metrics else summary.get("total_update_duration_s"),
            "point_total_wall_s": summary.get("point_total_wall_s"),
            "training_games_per_second_e2e": total_games / training_wall if training_wall else None,
            "rollout_games_per_second": total_games / rollout_wall if rollout_wall else None,
            "total_decision_samples": total_samples if metrics else summary.get("total_decision_samples"),
            "decisions_per_game": total_samples / total_games if metrics and total_games else summary.get("decisions_per_game"),
            "decision_samples_per_second": total_samples / training_wall if metrics and training_wall else None,
            "rl_workers_requested": summary.get("requested_workers", config.get("rl_workers")),
            "rl_workers_selected": summary.get("selected_workers"),
            "rl_workers_final": parallel.get("final_workers"),
            "rl_worker_candidates_tested": json.dumps(autotune.get("candidate_workers", [])),
            "autotune_iterations_per_test": autotune.get("iterations_per_test"),
            "autotune_games_per_test": autotune.get("games_per_test"),
            "autotune_reused_games": autotune.get("reused_game_count"),
            "autotune_duration_s": sum(
                float(item.get("elapsed_seconds", item.get("duration_s", 0.0)))
                for item in autotune.get("attempts", [])
            ),
            "autotune_discarded_games": autotune.get("discarded_game_count"),
            "fallback_count": parallel.get("fallback_count"),
            "safety_capped": parallel.get("safety_capped"),
            "min_available_memory_mb": parallel.get("min_available_memory_mb"),
            "peak_worker_rss_mb": parallel.get("peak_worker_rss_mb"),
            "peak_total_children_rss_mb": parallel.get("peak_total_children_rss_mb"),
            "pool_snapshot_count": summary.get("pool_snapshot_count", metrics[-1].get("pool_size") if metrics else None),
            "final_diagnostic_games": 2 * int(config.get("diagnostic_games", experiment_config.get("diagnostic_games", 0))),
            "mean_grad_norm": _mean(grad_norms),
            "std_grad_norm": _std(grad_norms),
            "p95_grad_norm": _percentile(grad_norms, 0.95),
            "clipped_iteration_count": clipped_count if metrics else summary.get("clipped_iteration_count"),
            "clipped_iteration_rate": clipped_count / len(metrics) if metrics else summary.get("clipped_iteration_rate"),
            "mean_entropy": _mean(entropies),
            "first_entropy": entropies[0] if entropies else None,
            "final_entropy": entropies[-1] if entropies else None,
            "mean_reward": _mean(reward_means),
            "mean_reward_std": _mean(reward_stds),
            "first_training_win_rate": metrics[0].get("batch_win_rate") if metrics else None,
            "final_training_win_rate": metrics[-1].get("batch_win_rate") if metrics else None,
            "final_training_moving_average_win_rate": metrics[-1].get("moving_average_win_rate") if metrics else None,
            "mean_value_loss": _mean(value_losses),
            "final_value_loss": value_losses[-1] if value_losses else None,
            "ppo_optimizer_steps": sum(
                int(item.get("optimizer_steps", 0)) for item in ppo_rows
            ),
            "ppo_mean_epochs": _mean(
                [item.get("epochs_completed") for item in ppo_rows]
            ),
            "ppo_kl_stop_count": sum(
                bool(item.get("stopped_by_kl")) for item in ppo_rows
            ),
            "ppo_mean_final_kl": _mean(
                [item.get("final_approx_kl") for item in ppo_rows]
            ),
            "ppo_max_kl": max(
                (float(item["max_approx_kl"]) for item in ppo_rows),
                default=None,
            ),
            "ppo_mean_clip_fraction": _mean(
                [item.get("final_clip_fraction") for item in ppo_rows]
            ),
            "ppo_mean_effective_minibatches": _mean(
                [item.get("effective_minibatches") for item in ppo_rows]
            ),
            "ppo_gpu_buffer_iterations": sum(
                item.get("buffer_location") == "gpu" for item in ppo_rows
            ),
            "ppo_ram_buffer_iterations": sum(
                item.get("buffer_location") != "gpu" for item in ppo_rows
            ),
            "failure_type": point.get("failure", {}).get("type"),
            "failure_message": point.get("failure", {}).get("message"),
            "failure_stage": point.get("failure", {}).get("stage"),
            "run_directory": str(run_dir),
        }
        row.update(_diagnostic_columns("heuristic", heuristic))
        row.update(_diagnostic_columns("random", random_summary))
        row["diagnostic_workers_heuristic_final"] = row.get("heuristic_workers_final")
        row["diagnostic_workers_random_final"] = row.get("random_workers_final")
        rows.append(row)

    for critic, hashes in sl_hashes.items():
        hashes.discard("None")
        if len(hashes) > 1:
            raise ValueError(
                f"Cannot aggregate critic={critic}: multiple SL hashes were found: {sorted(hashes)}"
            )
    return rows, warnings


def _statistics_columns(rows: list[dict[str, Any]], source: str, prefix: str) -> dict[str, Any]:
    values = [row.get(source) for row in rows if row.get(source) is not None]
    result = {
        f"{prefix}_mean": _mean(values),
        f"{prefix}_median": _median(values),
        f"{prefix}_std": _std(values),
    }
    if values:
        result[f"{prefix}_min"] = min(values)
        result[f"{prefix}_max"] = max(values)
        result[f"{prefix}_se"] = (_std(values) or 0.0) / math.sqrt(len(values))
    else:
        result[f"{prefix}_min"] = None
        result[f"{prefix}_max"] = None
        result[f"{prefix}_se"] = None
    return result


def pareto_flags(rows: list[dict[str, Any]]) -> dict[tuple[str, int], bool]:
    """Mark non-dominated quality/throughput aggregate points by critic."""
    flags: dict[tuple[str, int], bool] = {}
    for row in rows:
        quality = row.get("heuristic_win_rate_mean")
        speed = row.get("training_gps_median")
        efficient = quality is not None and speed is not None
        if efficient:
            for other in rows:
                if other is row or other["critic"] != row["critic"]:
                    continue
                other_quality = other.get("heuristic_win_rate_mean")
                other_speed = other.get("training_gps_median")
                if other_quality is None or other_speed is None:
                    continue
                if (
                    other_quality >= quality
                    and other_speed >= speed
                    and (other_quality > quality or other_speed > speed)
                ):
                    efficient = False
                    break
        flags[(row["critic"], int(row["games_per_iteration"]))] = efficient
    return flags


def aggregate_runs(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Aggregate completed models across seeds without pooling their games."""
    groups: dict[tuple[str, int], list[dict[str, Any]]] = {}
    planned: dict[tuple[str, int], int] = {}
    for row in rows:
        key = (str(row["critic"]), int(row["games_per_iteration"]))
        planned[key] = planned.get(key, 0) + 1
        if row.get("status") == "complete":
            groups.setdefault(key, []).append(row)
    aggregates: list[dict[str, Any]] = []
    for key in sorted(planned):
        completed = groups.get(key, [])
        critic, gpi = key
        base: dict[str, Any] = {
            "critic": critic,
            "games_per_iteration": gpi,
            "completed_runs": len(completed),
            "planned_runs": planned[key],
            "iterations": completed[0].get("iterations") if completed else None,
            "total_training_games": completed[0].get("total_training_games") if completed else None,
            "selected_workers_mode": None,
            "selected_workers_min": None,
            "selected_workers_max": None,
            "pareto_efficient": False,
            "quality_rank": None,
            "throughput_rank": None,
            "recommendation_status": "incomplete" if len(completed) < planned[key] else "not_selected",
        }
        if completed:
            workers = [row.get("rl_workers_selected") for row in completed if row.get("rl_workers_selected") is not None]
            if workers:
                base["selected_workers_mode"] = statistics.multimode(workers)[0]
                base["selected_workers_min"] = min(workers)
                base["selected_workers_max"] = max(workers)
        mappings = (
            ("training_wall_s", "training_wall"),
            ("training_games_per_second_e2e", "training_gps"),
            ("rollout_games_per_second", "rollout_gps"),
            ("decisions_per_game", "decisions_per_game"),
            ("clipped_iteration_rate", "clipped_rate"),
            ("mean_grad_norm", "grad_norm"),
            ("std_grad_norm", "within_run_grad_std"),
            ("mean_entropy", "entropy"),
            ("heuristic_win_rate", "heuristic_win_rate"),
            ("heuristic_score_rate", "heuristic_score_rate"),
            ("random_win_rate", "random_win_rate"),
            ("random_score_rate", "random_score_rate"),
            ("heuristic_mean_turns", "heuristic_mean_turns"),
            ("heuristic_mean_agent_remaining_pips", "heuristic_agent_pips"),
            ("heuristic_mean_opponent_remaining_pips", "heuristic_opponent_pips"),
        )
        for source, prefix in mappings:
            base.update(_statistics_columns(completed, source, prefix))
        aggregates.append(base)

    flags = pareto_flags(aggregates)
    for critic in sorted({row["critic"] for row in aggregates}):
        cohort = [row for row in aggregates if row["critic"] == critic]
        quality = sorted(
            (row for row in cohort if row.get("heuristic_win_rate_mean") is not None),
            key=lambda row: (-row["heuristic_win_rate_mean"], row["games_per_iteration"]),
        )
        throughput = sorted(
            (row for row in cohort if row.get("training_gps_median") is not None),
            key=lambda row: (-row["training_gps_median"], row["games_per_iteration"]),
        )
        for rank, row in enumerate(quality, 1):
            row["quality_rank"] = rank
        for rank, row in enumerate(throughput, 1):
            row["throughput_rank"] = rank
    for row in aggregates:
        row["pareto_efficient"] = flags[(row["critic"], row["games_per_iteration"])]
    return aggregates


def recommend_configurations(aggregates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Apply quality, efficiency, Pareto, and one-standard-error views."""
    rankings: list[dict[str, Any]] = []
    for critic in sorted({row["critic"] for row in aggregates}):
        cohort = [
            row for row in aggregates
            if row["critic"] == critic and row.get("heuristic_win_rate_mean") is not None
        ]
        if not cohort:
            continue
        best = max(cohort, key=lambda row: row["heuristic_win_rate_mean"])
        threshold = best["heuristic_win_rate_mean"] - (best.get("heuristic_win_rate_se") or 0.0)
        eligible = [row for row in cohort if row["heuristic_win_rate_mean"] >= threshold]
        recommended = max(
            eligible,
            key=lambda row: (row.get("training_gps_median") or -math.inf, -row["games_per_iteration"]),
        )
        score_best = max(cohort, key=lambda row: row.get("heuristic_score_rate_mean") or -math.inf)
        score_threshold = score_best.get("heuristic_score_rate_mean") - (score_best.get("heuristic_score_rate_se") or 0.0)
        score_eligible = [row for row in cohort if (row.get("heuristic_score_rate_mean") or -math.inf) >= score_threshold]
        score_recommended = max(
            score_eligible,
            key=lambda row: (row.get("training_gps_median") or -math.inf, -row["games_per_iteration"]),
        )
        max_seeds = max(row["completed_runs"] for row in cohort)
        evidence = "exploratory" if max_seeds < 3 else "multi-seed"
        for row in cohort:
            selected = row is recommended
            if selected:
                row["recommendation_status"] = f"recommended_{evidence}"
            rankings.append({
                "critic": critic,
                "games_per_iteration": row["games_per_iteration"],
                "quality_rank": row.get("quality_rank"),
                "throughput_rank": row.get("throughput_rank"),
                "pareto_efficient": row.get("pareto_efficient"),
                "one_se_eligible": row in eligible,
                "recommended": selected,
                "evidence_status": evidence,
                "best_quality_gpi": best["games_per_iteration"],
                "best_efficiency_gpi": max(cohort, key=lambda item: item.get("training_gps_median") or -math.inf)["games_per_iteration"],
                "one_se_threshold": threshold,
                "score_rate_recommended_gpi": score_recommended["games_per_iteration"],
                "score_rate_changes_recommendation": score_recommended["games_per_iteration"] != recommended["games_per_iteration"],
                "rule": "highest throughput among GPIs within one seed-level standard error of the best mean heuristic win rate",
            })
    return rankings


def build_pairwise_deltas(aggregates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for critic in sorted({row["critic"] for row in aggregates}):
        cohort = [row for row in aggregates if row["critic"] == critic]
        baseline = next((row for row in cohort if row["games_per_iteration"] == 40), None)
        valid = [row for row in cohort if row.get("heuristic_win_rate_mean") is not None]
        best = max(valid, key=lambda row: row["heuristic_win_rate_mean"]) if valid else None
        for row in cohort:
            rows.append({
                "critic": critic,
                "games_per_iteration": row["games_per_iteration"],
                "heuristic_win_rate_delta_vs_gpi40": (
                    row.get("heuristic_win_rate_mean") - baseline.get("heuristic_win_rate_mean")
                    if baseline and row.get("heuristic_win_rate_mean") is not None and baseline.get("heuristic_win_rate_mean") is not None else None
                ),
                "heuristic_score_rate_delta_vs_gpi40": (
                    row.get("heuristic_score_rate_mean") - baseline.get("heuristic_score_rate_mean")
                    if baseline and row.get("heuristic_score_rate_mean") is not None and baseline.get("heuristic_score_rate_mean") is not None else None
                ),
                "heuristic_win_rate_delta_vs_best": (
                    row.get("heuristic_win_rate_mean") - best.get("heuristic_win_rate_mean")
                    if best and row.get("heuristic_win_rate_mean") is not None else None
                ),
                "best_quality_gpi": best.get("games_per_iteration") if best else None,
            })
    return rows


def build_autotune_rows(experiment_dir: Path, run_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for run in run_rows:
        summary = _load_json(
            experiment_dir / "runs" / run["run_key"] / "training_summary.json"
        ) or {}
        autotune = summary.get("autotune", {})
        for index, attempt in enumerate(autotune.get("attempts", []), 1):
            rows.append({
                "run_key": run["run_key"],
                "critic": run["critic"],
                "seed": run["seed"],
                "games_per_iteration": run["games_per_iteration"],
                "attempt": index,
                "workers": attempt.get("requested_workers"),
                "duration_s": attempt.get("elapsed_seconds", attempt.get("duration_s")),
                "games_per_second": attempt.get("games_per_second"),
                "marginal_gain": attempt.get("improvement_over_previous"),
                "passed": attempt.get("success", attempt.get("passed")),
                "failure_reason": attempt.get("failure", attempt.get("failure_reason")),
                "planned_games": attempt.get("planned_games"),
                "reused_games": 0 if "success" in attempt else attempt.get("completed_games"),
                "discarded_games": (
                    attempt.get("actual_games") if "success" in attempt else 0
                ),
            })
    return rows


def _write_plots(output_dir: Path, aggregates: list[dict[str, Any]], run_rows: list[dict[str, Any]]) -> list[str]:
    if not aggregates:
        return []
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plot_dir = output_dir / "plots"
    plot_dir.mkdir(parents=True, exist_ok=True)
    written: list[str] = []

    def line_plot(filename: str, fields: list[tuple[str, str]], title: str, ylabel: str) -> None:
        fig, ax = plt.subplots(figsize=(8, 4.8), dpi=150)
        for critic in sorted({row["critic"] for row in aggregates}):
            cohort = sorted((row for row in aggregates if row["critic"] == critic), key=lambda row: row["games_per_iteration"])
            for field, label in fields:
                error_field = field.replace("_mean", "_se") if field.endswith("_mean") else None
                points = [
                    (
                        row["games_per_iteration"],
                        row.get(field),
                        row.get(error_field) if error_field else None,
                    )
                    for row in cohort
                    if row.get(field) is not None
                ]
                if not points:
                    continue
                x, y, errors = zip(*points)
                ax.errorbar(range(len(x)), y, yerr=[error or 0.0 for error in errors], marker="o", label=f"critic {critic} - {label}")
                ax.set_xticks(range(len(x)), [str(value) for value in x])
        ax.set_xlabel("Games per iteration (categorical positions; labels are actual GPI)")
        ax.set_ylabel(ylabel)
        ax.set_title(title)
        ax.grid(axis="y", alpha=0.25)
        ax.legend(frameon=False)
        fig.tight_layout()
        path = plot_dir / filename
        fig.savefig(path)
        plt.close(fig)
        written.append(str(path))

    line_plot("heuristic_win_rate_vs_gpi.png", [("heuristic_win_rate_mean", "win rate")], "RL quality against heuristic", "Win rate")
    line_plot("random_win_rate_vs_gpi.png", [("random_win_rate_mean", "win rate")], "RL quality against random", "Win rate")
    line_plot("score_rate_vs_gpi.png", [("heuristic_score_rate_mean", "heuristic"), ("random_score_rate_mean", "random")], "Score rate by GPI", "Score rate")
    line_plot("training_throughput_vs_gpi.png", [("training_gps_mean", "end-to-end"), ("rollout_gps_mean", "rollout")], "Training throughput", "Training games/s")
    line_plot("training_time_vs_gpi.png", [("training_wall_mean", "wall time")], "Training wall time", "Seconds")
    line_plot("workers_vs_gpi.png", [("selected_workers_mode", "selected workers")], "Selected rollout workers", "Workers")
    line_plot("decisions_per_game_vs_gpi.png", [("decisions_per_game_mean", "decisions/game")], "Real decisions per training game", "Decision samples/game")
    line_plot(
        "gradient_noise_vs_gpi.png",
        [("grad_norm_mean", "mean gradient norm"), ("within_run_grad_std_mean", "within-run gradient std"), ("clipped_rate_mean", "clipped rate")],
        "Gradient magnitude, variability, and clipping by GPI",
        "Gradient metric",
    )

    fig, ax = plt.subplots(figsize=(7.5, 5), dpi=150)
    for critic in sorted({row["critic"] for row in aggregates}):
        cohort = [row for row in aggregates if row["critic"] == critic and row.get("training_gps_median") is not None and row.get("heuristic_win_rate_mean") is not None]
        ax.scatter([row["training_gps_median"] for row in cohort], [row["heuristic_win_rate_mean"] for row in cohort], label=f"critic {critic}")
        for row in cohort:
            ax.annotate(str(row["games_per_iteration"]), (row["training_gps_median"], row["heuristic_win_rate_mean"]), xytext=(4, 3), textcoords="offset points", fontsize=8)
    ax.set_xlabel("Median end-to-end training games/s")
    ax.set_ylabel("Mean win rate vs heuristic")
    ax.set_title("Quality-throughput Pareto view")
    ax.grid(alpha=0.25)
    ax.legend(frameon=False)
    fig.tight_layout()
    pareto_path = plot_dir / "quality_throughput_pareto.png"
    fig.savefig(pareto_path)
    plt.close(fig)
    written.append(str(pareto_path))

    fig, ax = plt.subplots(figsize=(8, 5), dpi=150)
    for run in run_rows:
        metrics = _load_metrics(Path(run["run_directory"]) / "training_metrics.jsonl")
        if not metrics:
            continue
        ax.plot([item["cumulative_training_games"] for item in metrics], [item["moving_average_win_rate"] for item in metrics], alpha=0.5, label=f"{run['critic']} gpi={run['games_per_iteration']} s={run['seed']}")
    ax.set_xlabel("Cumulative training games")
    ax.set_ylabel("Moving-average training win rate")
    ax.set_title("On-policy training curves (not final strength estimates)")
    ax.grid(alpha=0.25)
    if len(run_rows) <= 20:
        ax.legend(frameon=False, fontsize=7)
    fig.tight_layout()
    curve_path = plot_dir / "learning_curves_vs_training_games.png"
    fig.savefig(curve_path)
    plt.close(fig)
    written.append(str(curve_path))
    return written


def _write_workbook(
    path: Path,
    sheets: list[tuple[str, list[dict[str, Any]]]],
    sheet_fields: dict[str, Iterable[str]] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX output. Install it with "
            "python -m pip install -r train_script/requirements_gpi_sweep.txt "
            "or use --csv-only."
        ) from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    percent_fragments = ("_rate", "_pct", "clipped_iteration_rate", "marginal_gain")
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        fields = _fieldnames(rows) or list((sheet_fields or {}).get(title, ("status",)))
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field) for field in fields])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="355C7D")
        for column_index, field in enumerate(fields, 1):
            values = [str(field)] + [str(row.get(field, "")) for row in rows[:500]]
            sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = min(45, max(10, max(map(len, values)) + 2))
            if any(fragment in field for fragment in percent_fragments):
                for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                    for item in cell:
                        if isinstance(item.value, (int, float)):
                            item.number_format = "0.00%"
            if field in {"status", "recommendation_status"}:
                for row_index in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row_index, column_index)
                    if cell.value in {"failed", "incomplete"}:
                        cell.fill = PatternFill("solid", fgColor="F4CCCC")
            if field in {"model_path", "run_directory"}:
                for row_index in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row_index, column_index)
                    if isinstance(cell.value, str) and cell.value:
                        target = Path(cell.value)
                        cell.hyperlink = os.path.relpath(target, path.parent)
                        cell.style = "Hyperlink"
            if "win_rate" in field or field.endswith("_gps_mean") or field.endswith("_gps_median"):
                letter = sheet.cell(1, column_index).column_letter
                if sheet.max_row >= 2:
                    sheet.conditional_formatting.add(
                        f"{letter}2:{letter}{sheet.max_row}",
                        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
                    )
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.tmp-{os.getpid()}.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _configuration_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for section in ("configuration", "environment", "budget", "sl_checkpoint"):
        for key, value in (manifest.get(section) or {}).items():
            rows.append({"section": section, "name": key, "value": json.dumps(value) if isinstance(value, (dict, list)) else value})
    rows.append({"section": "methodology", "name": "controlled_fields", "value": "training games, SL checkpoint, pool-refresh frequency, hyperparameters, diagnostics"})
    rows.append({"section": "methodology", "name": "deliberately_changed_fields", "value": "games per iteration, update count"})
    return rows


def _dictionary_rows(fields: Iterable[str]) -> list[dict[str, Any]]:
    rows = []
    for field in fields:
        unit = "fraction" if "rate" in field or field.endswith("_pct") else "seconds" if field.endswith("_s") else "count" if any(token in field for token in ("games", "iterations", "count", "workers")) else "value"
        rows.append({"column": field, "unit": unit, "type": "numeric or text", "meaning": field.replace("_", " ")})
    return rows


def build_report(
    experiment_dir: Path | str,
    output_dir: Path | str | None = None,
    *,
    csv_only: bool = False,
    generate_plots: bool = True,
) -> dict[str, Any]:
    """Validate disk artifacts and atomically publish the complete report."""
    experiment_dir = Path(experiment_dir).resolve()
    output_dir = Path(output_dir or experiment_dir / "report").resolve()
    manifest = _load_json(experiment_dir / "experiment_manifest.json", required=True)
    runs, warnings = build_run_rows(experiment_dir)
    aggregates = aggregate_runs(runs)
    rankings = recommend_configurations(aggregates)
    pairwise_deltas = build_pairwise_deltas(aggregates)
    autotune = build_autotune_rows(experiment_dir, runs)
    incomplete = [row["run_key"] for row in runs if row.get("status") != "complete"]
    recommended = [row for row in rankings if row.get("recommended")]
    report = {
        "schema_version": SCHEMA_VERSION,
        "experiment_metadata": {
            "experiment_id": manifest.get("experiment_id"),
            "experiment_dir": str(experiment_dir),
            "configuration_fingerprint": manifest.get("configuration_fingerprint"),
        },
        "configurations": manifest.get("configuration"),
        "runs": runs,
        "aggregates": aggregates,
        "rankings": rankings,
        "warnings": warnings,
        "incomplete_runs": incomplete,
        "recommended_configuration": recommended,
        "recommendation_rule": "highest median throughput within one seed-level standard error of the best mean heuristic win rate; exploratory below three seeds",
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    tables = {
        "runs": runs,
        "aggregates": aggregates,
        "rankings": rankings,
        "autotune": autotune,
        "pairwise_deltas": pairwise_deltas,
    }
    explicit_fields = {
        "rankings": RANKING_FIELDS,
        "autotune": AUTOTUNE_FIELDS,
        "pairwise_deltas": PAIRWISE_DELTA_FIELDS,
    }
    for name, rows in tables.items():
        _atomic_csv(output_dir / REPORT_CSV_FILES[name], rows, explicit_fields.get(name))
    configuration_rows = _configuration_rows(manifest)
    dictionary_rows = _dictionary_rows(_fieldnames(runs) + _fieldnames(aggregates))
    _atomic_csv(output_dir / "gpi_sweep_configuration.csv", configuration_rows)
    _atomic_csv(output_dir / "gpi_sweep_data_dictionary.csv", dictionary_rows)
    _atomic_json(output_dir / "gpi_sweep_report.json", report)
    plots = _write_plots(output_dir, aggregates, runs) if generate_plots else []
    if not csv_only:
        _write_workbook(
            output_dir / "gpi_sweep_results.xlsx",
            [
                ("Runs", runs),
                ("Aggregate", aggregates),
                ("Ranking", rankings),
                ("Autotune", autotune),
                ("Configuration", configuration_rows),
                ("Data_dictionary", dictionary_rows),
                ("Pairwise_deltas", pairwise_deltas),
            ],
            sheet_fields={
                "Ranking": RANKING_FIELDS,
                "Autotune": AUTOTUNE_FIELDS,
                "Pairwise_deltas": PAIRWISE_DELTA_FIELDS,
            },
        )
    report["output_dir"] = str(output_dir)
    report["plots"] = plots
    return report


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rebuild the RL games-per-iteration sweep report entirely from disk artifacts.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("experiment_dir", type=Path, help="Experiment results directory containing experiment_manifest.json.")
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--csv-only", action="store_true", help="Skip XLSX output and do not import openpyxl.")
    parser.add_argument("--no-plots", action="store_true", help="Skip comparative PNG plots.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    report = build_report(
        args.experiment_dir,
        args.output_dir,
        csv_only=args.csv_only,
        generate_plots=not args.no_plots,
    )
    print(f"GPI sweep report written to {report['output_dir']}")
    if report["incomplete_runs"]:
        print(f"Incomplete runs: {len(report['incomplete_runs'])}")


if __name__ == "__main__":
    main()
