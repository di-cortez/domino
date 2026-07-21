"""Tests for the controlled RL games-per-iteration sweep."""

from __future__ import annotations

import json
import os
from pathlib import Path
import tempfile
import unittest
from unittest import mock

import numpy as np

from diagnostics.rl_gpi_sweep_report import (
    _write_workbook,
    aggregate_runs,
    build_pairwise_deltas,
    pareto_flags,
    recommend_configurations,
)
from train_script import run_rl_games_per_iteration_sweep as sweep
from training import self_play


def base_config(**overrides):
    config = {
        "total_training_games": 384_000,
        "gpi_values": sweep.DEFAULT_GPI_VALUES,
        "seeds": sweep.DEFAULT_SEEDS,
        "critic_mode": "off",
        "diagnostic_games": 10_000,
        "checkpoint_count": 10,
        "checkpoint_evaluation_games": 200,
        "rl_workers": "auto",
        "diagnostic_workers": 1,
    }
    config.update(overrides)
    return config


class PlanningTests(unittest.TestCase):
    def test_standard_iterations_match_exact_table(self):
        plan = sweep.build_run_plan(base_config(seeds=(42,)))
        actual = {
            row["games_per_iteration"]: row["iterations"]
            for row in plan
        }
        self.assertEqual(actual, {
            40: 9600,
            80: 4800,
            160: 2400,
            320: 1200,
            640: 600,
            960: 400,
            1280: 300,
        })
        self.assertTrue(all(
            row["iterations"] * row["games_per_iteration"] == 384_000
            for row in plan
        ))

    def test_non_divisible_total_fails_before_training(self):
        with self.assertRaisesRegex(ValueError, "not exactly divisible"):
            sweep.build_run_plan(base_config(total_training_games=100, gpi_values=(40,)))

    def test_duplicate_and_nonpositive_values_fail(self):
        for overrides in (
            {"gpi_values": (40, 40)},
            {"gpi_values": (40, 0)},
            {"seeds": (42, 42)},
        ):
            with self.subTest(overrides=overrides), self.assertRaises(ValueError):
                sweep.build_run_plan(base_config(**overrides))

    def test_checkpoint_count_is_identical_for_every_point(self):
        plan = sweep.build_run_plan(base_config(seeds=(42,)))
        intervals = {
            row["games_per_iteration"]: row["iterations"] // 10
            for row in plan
        }
        self.assertTrue(all(row["iterations"] % intervals[row["games_per_iteration"]] == 0 for row in plan))
        self.assertEqual({row["iterations"] // intervals[row["games_per_iteration"]] for row in plan}, {10})

    def test_plan_order_is_deterministic_and_names_do_not_collide(self):
        first = sweep.build_run_plan(base_config(critic_mode="both"))
        second = sweep.build_run_plan(base_config(critic_mode="both"))
        self.assertEqual(first, second)
        keys = [point["run_key"] for point in first]
        self.assertEqual(len(keys), len(set(keys)))
        self.assertIn("critic_on_gpi", " ".join(keys))
        self.assertIn("games0384000_seed42", " ".join(keys))

    def test_diagnostic_seeds_match_across_gpis_but_not_opponents(self):
        heuristic = sweep.stable_diagnostic_seed(42, "heuristic")
        self.assertEqual(heuristic, sweep.stable_diagnostic_seed(42, "heuristic"))
        self.assertNotEqual(heuristic, sweep.stable_diagnostic_seed(42, "random"))

    @unittest.skipUnless(sweep.DEFAULT_SL_WEIGHTS_PATH.is_file(), "SL smoke checkpoint is unavailable")
    def test_dry_run_has_no_output_or_training_side_effect(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            args = sweep.parse_args([
                "--total-training-games", "80",
                "--games-per-iteration-values", "40", "80",
                "--seeds", "42",
                "--checkpoint-count", "1",
                "--diagnostic-games", "20",
                "--results-dir", str(root / "results"),
                "--model-dir", str(root / "models"),
                "--csv-only",
                "--dry-run",
            ])
            with mock.patch.object(self_play, "train") as train:
                result = sweep.run_experiment(args)
            train.assert_not_called()
            self.assertEqual(len(result["plan"]), 2)
            self.assertFalse((root / "results").exists())
            self.assertFalse((root / "models").exists())


class MetricCallbackTests(unittest.TestCase):
    @unittest.skipUnless(sweep.DEFAULT_SL_WEIGHTS_PATH.is_file(), "SL smoke checkpoint is unavailable")
    def test_callback_receives_one_json_serializable_row_per_iteration(self):
        with tempfile.TemporaryDirectory() as temporary:
            rows = []
            summary = self_play.train(
                iterations=2,
                games_per_iteration=2,
                checkpoint_interval=2,
                evaluation_games=2,
                pool_interval=1,
                max_pool_size=2,
                sl_weights_path=str(sweep.DEFAULT_SL_WEIGHTS_PATH),
                rl_weights_path=str(Path(temporary) / "model.npz"),
                seed=101,
                device="cpu",
                workers=1,
                quiet=True,
                numbered_checkpoints=True,
                metrics_callback=rows.append,
            )
            self.assertEqual([row["iteration"] for row in rows], [1, 2])
            json.dumps(rows)
            self.assertEqual(summary["total_training_games"], 4)
            self.assertEqual(summary["total_decision_samples"], sum(row["decision_sample_count"] for row in rows))
            self.assertTrue(rows[-1]["checkpoint_written"])

    @unittest.skipUnless(sweep.DEFAULT_SL_WEIGHTS_PATH.is_file(), "SL smoke checkpoint is unavailable")
    def test_none_callback_preserves_seeded_weights(self):
        with tempfile.TemporaryDirectory() as temporary:
            paths = [Path(temporary) / "without.npz", Path(temporary) / "with.npz"]
            rows = []
            common = dict(
                iterations=2,
                games_per_iteration=2,
                checkpoint_interval=2,
                evaluation_games=2,
                pool_interval=1,
                max_pool_size=2,
                sl_weights_path=str(sweep.DEFAULT_SL_WEIGHTS_PATH),
                seed=202,
                device="cpu",
                workers=1,
                quiet=True,
            )
            self_play.train(rl_weights_path=str(paths[0]), metrics_callback=None, **common)
            self_play.train(rl_weights_path=str(paths[1]), metrics_callback=rows.append, **common)
            with np.load(paths[0]) as left, np.load(paths[1]) as right:
                for name in left.files:
                    np.testing.assert_array_equal(left[name], right[name])

    def test_resume_truncates_metrics_after_checkpoint(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "metrics.jsonl"
            path.write_text("".join(json.dumps({"iteration": value}) + "\n" for value in range(1, 5)), encoding="utf-8")
            retained = sweep.truncate_metrics_file(path, 2)
            self.assertEqual([row["iteration"] for row in retained], [1, 2])
            self.assertEqual([json.loads(line)["iteration"] for line in path.read_text().splitlines()], [1, 2])

    def test_weight_without_resume_state_is_ignored(self):
        with tempfile.TemporaryDirectory() as temporary:
            base = Path(temporary) / "model.npz"
            (Path(temporary) / "model_iter000010.npz").write_bytes(b"not enough")
            self.assertIsNone(sweep.find_latest_resume_pair(base, 10))

    def test_elapsed_metric_segments_are_accumulated_across_resume(self):
        metrics = [
            {"elapsed_training_s": 1.0},
            {"elapsed_training_s": 2.5},
            {"elapsed_training_s": 0.75},
            {"elapsed_training_s": 1.5},
        ]
        self.assertEqual(sweep._metrics_elapsed_s(metrics), 4.0)

    def test_atomic_failure_preserves_previous_file(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "report.json"
            path.write_text("valid", encoding="utf-8")
            with mock.patch.object(sweep.os, "replace", side_effect=OSError("simulated")):
                with self.assertRaisesRegex(OSError, "simulated"):
                    sweep.atomic_write_text(path, "replacement")
            self.assertEqual(path.read_text(encoding="utf-8"), "valid")

    @unittest.skipUnless(sweep.DEFAULT_SL_WEIGHTS_PATH.is_file(), "SL smoke checkpoint is unavailable")
    def test_complete_numbered_pair_is_discovered(self):
        with tempfile.TemporaryDirectory() as temporary:
            base = Path(temporary) / "model.npz"
            self_play.train(
                iterations=1,
                games_per_iteration=2,
                checkpoint_interval=1,
                evaluation_games=2,
                pool_interval=1,
                max_pool_size=1,
                sl_weights_path=str(sweep.DEFAULT_SL_WEIGHTS_PATH),
                rl_weights_path=str(base),
                seed=303,
                device="cpu",
                workers=1,
                quiet=True,
                numbered_checkpoints=True,
            )
            pair = sweep.find_latest_resume_pair(base, 1)
            self.assertIsNotNone(pair)
            self.assertEqual(pair[2]["completed_iteration"], 1)

    @unittest.skipUnless(sweep.DEFAULT_SL_WEIGHTS_PATH.is_file(), "SL smoke checkpoint is unavailable")
    def test_csv_only_dry_run_does_not_import_openpyxl(self):
        with tempfile.TemporaryDirectory() as temporary:
            original_import = __import__

            def guarded_import(name, *args, **kwargs):
                if name == "openpyxl" or name.startswith("openpyxl."):
                    raise AssertionError("openpyxl must not be imported in csv-only mode")
                return original_import(name, *args, **kwargs)

            args = sweep.parse_args([
                "--total-training-games", "80",
                "--games-per-iteration-values", "40", "80",
                "--seeds", "42",
                "--checkpoint-count", "1",
                "--diagnostic-games", "20",
                "--results-dir", str(Path(temporary) / "results"),
                "--model-dir", str(Path(temporary) / "models"),
                "--csv-only",
                "--dry-run",
            ])
            with mock.patch("builtins.__import__", side_effect=guarded_import):
                sweep.run_experiment(args)


class ReportMathTests(unittest.TestCase):
    def synthetic_runs(self):
        rows = []
        for gpi, qualities, speeds in (
            (40, (0.60, 0.62, 0.61), (100, 105, 95)),
            (80, (0.603, 0.613, 0.608), (160, 170, 165)),
            (160, (0.50, 0.51, 0.49), (220, 230, 225)),
        ):
            for seed, (quality, speed) in enumerate(zip(qualities, speeds), 42):
                rows.append({
                    "critic": "off",
                    "games_per_iteration": gpi,
                    "status": "complete",
                    "iterations": 384_000 // gpi,
                    "total_training_games": 384_000,
                    "rl_workers_selected": 2,
                    "training_wall_s": 384_000 / speed,
                    "training_games_per_second_e2e": speed,
                    "rollout_games_per_second": speed * 1.2,
                    "decisions_per_game": 7.0,
                    "clipped_iteration_rate": 0.1,
                    "heuristic_win_rate": quality,
                    "heuristic_score_rate": quality,
                    "random_win_rate": 0.8,
                    "random_score_rate": 0.8,
                    "heuristic_mean_turns": 25,
                    "heuristic_mean_agent_remaining_pips": 5,
                    "heuristic_mean_opponent_remaining_pips": 8,
                    "seed": seed,
                })
        return rows

    def test_aggregation_pareto_and_one_se_rule(self):
        aggregates = aggregate_runs(self.synthetic_runs())
        self.assertEqual(len(aggregates), 3)
        flags = pareto_flags(aggregates)
        self.assertTrue(flags[("off", 40)])
        self.assertTrue(flags[("off", 80)])
        ranking = recommend_configurations(aggregates)
        recommended = [row for row in ranking if row["recommended"]]
        self.assertEqual(len(recommended), 1)
        self.assertEqual(recommended[0]["games_per_iteration"], 80)
        self.assertEqual(recommended[0]["evidence_status"], "multi-seed")

    def test_critic_cohorts_remain_separate(self):
        rows = self.synthetic_runs()
        rows.append({**rows[0], "critic": "on", "seed": 99})
        aggregates = aggregate_runs(rows)
        self.assertEqual({row["critic"] for row in aggregates}, {"off", "on"})

    def test_pairwise_deltas_use_gpi40_baseline(self):
        aggregates = aggregate_runs(self.synthetic_runs())
        deltas = build_pairwise_deltas(aggregates)
        baseline = next(row for row in deltas if row["games_per_iteration"] == 40)
        self.assertEqual(baseline["heuristic_win_rate_delta_vs_gpi40"], 0.0)

    def test_workbook_has_required_sheets_and_numeric_values(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "report.xlsx"
            sheets = [
                ("Runs", [{"status": "complete", "win_rate": 0.5}]),
                ("Aggregate", [{"games_per_iteration": 40}]),
                ("Ranking", [{"recommended": True}]),
                ("Autotune", []),
                ("Configuration", [{"name": "x", "value": 1}]),
                ("Data_dictionary", [{"column": "x"}]),
                ("Pairwise_deltas", [{"delta": 0.0}]),
            ]
            _write_workbook(path, sheets)
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=False, data_only=True)
            self.assertEqual(workbook.sheetnames, [name for name, _rows in sheets])
            self.assertEqual(workbook["Runs"].freeze_panes, "A2")
            self.assertEqual(workbook["Runs"]["B2"].value, 0.5)


class EndToEndSmokeTest(unittest.TestCase):
    @unittest.skipUnless(sweep.DEFAULT_SL_WEIGHTS_PATH.is_file(), "SL smoke checkpoint is unavailable")
    def test_real_minimal_sweep(self):
        """Two real points: 80 games each, 20 games per final opponent."""
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            args = sweep.parse_args([
                "--total-training-games", "80",
                "--games-per-iteration-values", "40", "80",
                "--seeds", "42",
                "--critic-mode", "off",
                "--diagnostic-games", "20",
                "--checkpoint-count", "1",
                "--checkpoint-evaluation-games", "2",
                "--device", "cpu",
                "--rl-workers", "1",
                "--diagnostic-workers", "1",
                "--diag-no-plots",
                "--quiet-training",
                "--run-id", "smoke",
                "--results-dir", str(root / "results"),
                "--model-dir", str(root / "models"),
                "--report-output-dir", str(root / "results" / "report"),
            ])
            report = sweep.run_experiment(args)
            self.assertEqual(len(report["runs"]), 2)
            self.assertTrue(all(row["status"] == "complete" for row in report["runs"]))
            self.assertTrue((root / "results" / "report" / "gpi_sweep_results.xlsx").is_file())
            self.assertTrue((root / "results" / "report" / "gpi_sweep_runs.csv").is_file())
            for row in report["runs"]:
                self.assertEqual(row["total_training_games"], 80)
                self.assertEqual(row["heuristic_games"], 20)
                self.assertEqual(row["random_games"], 20)


if __name__ == "__main__":
    unittest.main()
